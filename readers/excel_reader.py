"""
Excel Reader for ThreadBear

Handles .xlsx, .xls files using openpyxl with sheet-based segmentation.
"""
from pathlib import Path
from .registry import reader_registry
from .encoding import normalize_encoding


class ExcelReader:
    """Reader for Excel files."""

    @staticmethod
    def extract_text(path):
        """Read all sheets, return as markdown tables."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Excel reading requires openpyxl: pip install openpyxl")
        
        wb = openpyxl.load_workbook(path, data_only=True)
        parts = []
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"## Sheet: {sheet}\n")
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else '' for cell in row]
                if any(cells):  # Skip empty rows
                    rows.append(cells)
            
            if rows:
                # Format as markdown table
                header = rows[0]
                parts.append('| ' + ' | '.join(header) + ' |')
                parts.append('| ' + ' | '.join(['---'] * len(header)) + ' |')
                for row in rows[1:]:
                    padded = row[:len(header)] + [''] * max(0, len(header) - len(row))
                    parts.append('| ' + ' | '.join(str(cell) for cell in padded) + ' |')
                parts.append('')
        
        return '\n'.join(parts)

    @staticmethod
    def extract_segments(path):
        """Sheet-based segmentation. Each sheet = one segment."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Excel reading requires openpyxl: pip install openpyxl")
        
        wb = openpyxl.load_workbook(path, data_only=True)
        segments = []
        
        for i, sheet in enumerate(wb.sheetnames):
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else '' for cell in row]
                if any(cells):
                    rows.append(cells)
            
            if rows:
                # Format as markdown table
                header = rows[0]
                lines = ['| ' + ' | '.join(header) + ' |']
                lines.append('| ' + ' | '.join(['---'] * len(header)) + ' |')
                for row in rows[1:]:
                    padded = row[:len(header)] + [''] * max(0, len(header) - len(row))
                    lines.append('| ' + ' | '.join(str(cell) for cell in padded) + ' |')
                text = '\n'.join(lines)
                
                segments.append({
                    'text': text,
                    'start': i,
                    'end': i + 1,
                    'tokens': len(text) // 4,
                    'label': f'Sheet: {sheet}'
                })
        
        return segments


reader_registry.register(['.xlsx', '.xls'], ExcelReader, requires=['openpyxl'])
